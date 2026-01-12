#!/usr/bin/env python3
"""
Generate Excel spreadsheets from Library of Things JSON data.

Usage:
    python generate_xlsx.py --network minuteman
    python generate_xlsx.py --network cwmars
    python generate_xlsx.py --all
"""

import json
import argparse
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


# Paths
SCRIPT_DIR = Path(__file__).parent
DATA_DIR = SCRIPT_DIR.parent / "data"
OUTPUT_DIR = SCRIPT_DIR.parent / "outputs"


def load_json(filename: str) -> dict:
    """Load JSON data from data directory."""
    filepath = DATA_DIR / filename
    if not filepath.exists():
        raise FileNotFoundError(f"Data file not found: {filepath}")
    with open(filepath, 'r') as f:
        return json.load(f)


def create_xlsx(data: dict, output_name: str) -> str:
    """Create Excel spreadsheet from JSON data."""
    
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    highlight_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    items = data.get("items", [])
    libraries = data.get("libraries", [])
    network = data.get("network", {})
    
    # Sheet 1: All Items
    ws1 = wb.active
    ws1.title = "All Items"
    headers = ["Library", "Category", "Item Name", "Description"]
    ws1.append(headers)
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, item in enumerate(items, 2):
        ws1.append([
            item.get("library", ""),
            item.get("category", ""),
            item.get("name", ""),
            item.get("description", "")
        ])
        if row_num % 2 == 0:
            for col in range(1, 5):
                ws1.cell(row=row_num, column=col).fill = alt_fill
    
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 40
    ws1.column_dimensions['D'].width = 70
    
    # Sheet 2: By Category
    ws2 = wb.create_sheet("By Category")
    ws2.append(["Category", "Library", "Item Name", "Description"])
    
    for col in range(1, 5):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    sorted_items = sorted(items, key=lambda x: (x.get("category", ""), x.get("library", ""), x.get("name", "")))
    for item in sorted_items:
        ws2.append([
            item.get("category", ""),
            item.get("library", ""),
            item.get("name", ""),
            item.get("description", "")
        ])
    
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 30
    ws2.column_dimensions['C'].width = 40
    ws2.column_dimensions['D'].width = 70
    
    # Sheet 3: Home Improvement Focus
    ws3 = wb.create_sheet("Home & Tools")
    ws3.append(["Library", "Item Name", "Description", "Category"])
    
    tool_fill = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
    for col in range(1, 5):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = tool_fill
    
    home_categories = [
        "Home Improvement", "Measurement & Detection", "Home Inspection",
        "Auto/Vehicle", "Bicycle", "Gardening"
    ]
    home_items = [i for i in items if i.get("category") in home_categories]
    for item in home_items:
        ws3.append([
            item.get("library", ""),
            item.get("name", ""),
            item.get("description", ""),
            item.get("category", "")
        ])
    
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 70
    ws3.column_dimensions['D'].width = 22
    
    # Sheet 4: Library Directory
    ws4 = wb.create_sheet("Library Directory")
    ws4.append(["Library", "Location", "Phone", "Website", "Notes"])
    
    for col in range(1, 6):
        cell = ws4.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
    
    for lib in libraries:
        ws4.append([
            lib.get("name", ""),
            lib.get("location", ""),
            lib.get("phone", ""),
            lib.get("lot_url", lib.get("website", "")),
            lib.get("notes", "")
        ])
    
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 18
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 55
    ws4.column_dimensions['E'].width = 50
    
    # Sheet 5: Summary
    ws5 = wb.create_sheet("Summary")
    ws5.append([f"{network.get('name', 'Library')} - Library of Things Database"])
    ws5.merge_cells('A1:D1')
    ws5.cell(row=1, column=1).font = Font(bold=True, size=16)
    
    ws5.append([])
    ws5.append(["Generated:", datetime.now().strftime("%B %d, %Y")])
    ws5.append(["Total Items:", len(items)])
    ws5.append(["Libraries:", len(libraries)])
    ws5.append(["Network:", network.get("name", "Unknown")])
    ws5.append(["Region:", network.get("region", "Massachusetts")])
    ws5.append([])
    
    # Category counts
    ws5.append(["Items by Category:"])
    categories = {}
    for item in items:
        cat = item.get("category", "Unknown")
        categories[cat] = categories.get(cat, 0) + 1
    
    for cat, count in sorted(categories.items(), key=lambda x: -x[1]):
        ws5.append([f"  {cat}:", count])
    
    ws5.append([])
    ws5.append(["Items by Library:"])
    lib_counts = {}
    for item in items:
        lib = item.get("library", "Unknown")
        lib_counts[lib] = lib_counts.get(lib, 0) + 1
    
    for lib, count in sorted(lib_counts.items(), key=lambda x: -x[1]):
        ws5.append([f"  {lib}:", count])
    
    ws5.column_dimensions['A'].width = 50
    ws5.column_dimensions['B'].width = 15
    
    # Save
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / output_name
    wb.save(output_path)
    
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate Excel files from Library of Things data")
    parser.add_argument("--network", choices=["minuteman", "cwmars"], help="Network to generate")
    parser.add_argument("--all", action="store_true", help="Generate all networks")
    args = parser.parse_args()
    
    if args.all or args.network == "minuteman":
        try:
            data = load_json("minuteman_items.json")
            output = create_xlsx(data, "minuteman_library_of_things.xlsx")
            print(f"✓ Created: {output}")
            print(f"  Items: {len(data.get('items', []))}")
            print(f"  Libraries: {len(data.get('libraries', []))}")
        except Exception as e:
            print(f"✗ Error generating Minuteman: {e}")
    
    if args.all or args.network == "cwmars":
        try:
            data = load_json("cwmars_items.json")
            output = create_xlsx(data, "cwmars_library_of_things.xlsx")
            print(f"✓ Created: {output}")
            print(f"  Items: {len(data.get('items', []))}")
            print(f"  Libraries: {len(data.get('libraries', []))}")
        except Exception as e:
            print(f"✗ Error generating CWMARS: {e}")
    
    if not args.all and not args.network:
        parser.print_help()


if __name__ == "__main__":
    main()
